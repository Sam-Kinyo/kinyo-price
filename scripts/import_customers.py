"""
從 ERP 匯出的客戶查詢 xlsx 匯入 Firestore Customers collection
Doc ID = 客戶編號

Usage:
    python import_customers.py <path-to-xlsx> [--dry-run]
"""
import sys
import io
import os
import argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

# 欄位索引（從 0 起算）
COL_CODE = 1      # 客戶編號
COL_SHORT = 2     # 簡稱
COL_FULL = 3      # 全名
COL_TAX_ID = 10   # 統編
COL_PHONE = 11    # 電話
COL_ADDRESS = 12  # 地址


def clean(val):
    if val is None:
        return ''
    return str(val).strip()


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # header 在第 1 列 (index 1), 資料從 index 2 開始
    customers = []
    for i, row in enumerate(rows):
        if i < 2:
            continue
        code = clean(row[COL_CODE])
        if not code:
            continue
        customers.append({
            'code': code,
            'short': clean(row[COL_SHORT]),
            'full': clean(row[COL_FULL]),
            'taxId': clean(row[COL_TAX_ID]),
            'phone': clean(row[COL_PHONE]),
            'address': clean(row[COL_ADDRESS]),
        })
    return customers


def import_to_firestore(customers, cred_path, dry_run=False):
    if not dry_run:
        if not firebase_admin._apps:
            cred = credentials.Certificate(cred_path)
            firebase_admin.initialize_app(cred)
        db = firestore.client()

    batch_size = 400
    total = len(customers)
    written = 0

    for start in range(0, total, batch_size):
        batch_data = customers[start:start + batch_size]

        if dry_run:
            print(f'[DRY-RUN] 第 {start + 1} ~ {start + len(batch_data)} 筆 (共 {total})')
            for c in batch_data[:3]:
                print(f'  {c["code"]}: {c["short"]} / {c["full"]} / 統編 {c["taxId"]} / {c["phone"]} / {c["address"]}')
            continue

        batch = db.batch()
        for c in batch_data:
            doc_ref = db.collection('Customers').document(c['code'])
            batch.set(doc_ref, {
                'code': c['code'],
                'shortName': c['short'],
                'fullName': c['full'],
                'taxId': c['taxId'],
                'phone': c['phone'],
                'address': c['address'],
                'source': 'erp_xlsx',
                'updatedAt': firestore.SERVER_TIMESTAMP,
            }, merge=True)
        batch.commit()
        written += len(batch_data)
        print(f'  已寫入 {written}/{total}')

    return total, written


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', help='客戶查詢 xlsx 路徑')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--cred', default=r'D:\SAM-KINYO-WEBSITE\kinyo-price-autoimage\kinyo-price-admin.json')
    args = parser.parse_args()

    customers = parse_xlsx(args.xlsx)
    print(f'解析完成：共 {len(customers)} 筆')
    if customers:
        print('第一筆：', customers[0])
        print('最後一筆：', customers[-1])
    print()

    if args.dry_run:
        import_to_firestore(customers, args.cred, dry_run=True)
        print('\n[DRY-RUN] 完成，未寫入 Firestore。')
    else:
        total, written = import_to_firestore(customers, args.cred, dry_run=False)
        print(f'\n✅ 完成：共解析 {total} 筆，寫入 {written} 筆到 Customers collection')


if __name__ == '__main__':
    main()
