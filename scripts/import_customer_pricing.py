"""
從 ERP 客戶 C 表 xlsx 匯入該客戶的固定價
寫入 Firestore: Customers/{客戶編號}/Pricing/{型號} 子集合

規則:
- 第 1 列 (index 1) 取客戶代號
- 第 6 列 (index 6) 開始是資料
- 型號: 品名第一個空白前的 token，去掉非字母數字，轉大寫 (e.g. "AS-HP90 xxx" -> "ASHP90")
- 單價: 取「最後單價」欄
- 停產 (*停) 跳過
- 支援多個檔案一次匯入 (每個檔是一個客戶)

Usage:
    python import_customer_pricing.py <xlsx1> [<xlsx2> ...]
    python import_customer_pricing.py --dry-run <xlsx>
"""
import sys
import io
import os
import argparse
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore


CRED_PATH = r'D:\SAM-KINYO-WEBSITE\kinyo-price-autoimage\kinyo-price-admin.json'


def normalize_model(raw_product_name):
    """
    '  AS-HP90 分離式雙電壓旅行快煮壼 ' -> 'ASHP90'
    '  BP-080 多功能鴛鴦電火鍋 3L *停' -> ('BP080', True)  # 停產
    """
    if not raw_product_name:
        return None, False
    s = str(raw_product_name).strip()
    discontinued = '*停' in s
    # 取第一個空白前的 token
    first_token = s.split()[0] if s.split() else ''
    # 去掉非英數
    normalized = re.sub(r'[^A-Za-z0-9]', '', first_token).upper()
    return normalized or None, discontinued


def parse_c_table(path):
    """回傳 (customer_code, customer_name, [pricing_items])"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 7:
        raise ValueError(f'{path}: 格式不符 (列數太少)')

    # Row 1: 客戶代號 | OAZ041 | 電話 | 02-...
    customer_code = str(rows[1][2] or '').strip() if rows[1][2] else ''
    if not customer_code:
        raise ValueError(f'{path}: 找不到客戶代號 (應該在第 2 列 C 欄)')

    # Row 2: 客戶名稱 | xxxxx
    customer_name = str(rows[2][2] or '').strip() if rows[2][2] else ''

    items = []
    skipped_discontinued = 0
    skipped_empty = 0
    for i, row in enumerate(rows):
        if i < 6:
            continue
        # Row: [None, 品名, None, 數量小計, 金額小計, 最後單價, 最後交易日, 國際條碼, None]
        product_name = row[1]
        last_price = row[5]
        last_trans_date = row[6]
        total_qty = row[3]
        total_amount = row[4]

        if not product_name:
            continue

        model, discontinued = normalize_model(product_name)
        if not model:
            skipped_empty += 1
            continue
        if discontinued:
            skipped_discontinued += 1
            continue
        if last_price is None:
            continue

        items.append({
            'model': model,
            'unitPrice': int(float(last_price)),
            'lastOrderDate': last_trans_date.isoformat() if last_trans_date else None,
            'totalQty': int(float(total_qty)) if total_qty else 0,
            'totalAmount': int(float(total_amount)) if total_amount else 0,
            'rawProductName': str(product_name).strip(),
        })

    return {
        'customerCode': customer_code,
        'customerName': customer_name,
        'items': items,
        'skipped': {'discontinued': skipped_discontinued, 'empty': skipped_empty}
    }


def import_to_firestore(parsed, dry_run=False):
    code = parsed['customerCode']
    items = parsed['items']
    if not items:
        print(f'  {code}: 無資料可匯入')
        return 0

    if dry_run:
        print(f'\n[DRY-RUN] {code} {parsed["customerName"]}')
        print(f'  跳過: 停產 {parsed["skipped"]["discontinued"]} 筆, 空白 {parsed["skipped"]["empty"]} 筆')
        print(f'  將寫入 {len(items)} 筆:')
        for it in items[:5]:
            print(f'    {it["model"]}: ${it["unitPrice"]} (最後 {it["lastOrderDate"]}, 共買 {it["totalQty"]} 台 ${it["totalAmount"]})')
        if len(items) > 5:
            print(f'    ... 共 {len(items)} 筆')
        return len(items)

    if not firebase_admin._apps:
        cred = credentials.Certificate(CRED_PATH)
        firebase_admin.initialize_app(cred)
    db = firestore.client()

    # 檢查客戶是否存在
    cust_doc = db.collection('Customers').document(code).get()
    if not cust_doc.exists:
        print(f'  ⚠️ Customers/{code} 不存在，跳過整份')
        return 0

    # 批次寫入
    batch_size = 400
    written = 0
    for start in range(0, len(items), batch_size):
        chunk = items[start:start + batch_size]
        batch = db.batch()
        for it in chunk:
            ref = db.collection('Customers').document(code).collection('Pricing').document(it['model'])
            batch.set(ref, {
                'model': it['model'],
                'unitPrice': it['unitPrice'],
                'lastOrderDate': it['lastOrderDate'],
                'totalQty': it['totalQty'],
                'totalAmount': it['totalAmount'],
                'rawProductName': it['rawProductName'],
                'source': 'erp_c_table',
                'updatedAt': firestore.SERVER_TIMESTAMP,
            }, merge=True)
        batch.commit()
        written += len(chunk)

    print(f'  ✅ {code} {parsed["customerName"]}: {written} 筆 Pricing 寫入完成')
    return written


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('files', nargs='+', help='C表 xlsx 路徑 (可多個)')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    total = 0
    for path in args.files:
        print(f'\n處理 {path}...')
        try:
            parsed = parse_c_table(path)
            print(f'  客戶: {parsed["customerCode"]} {parsed["customerName"]}')
            print(f'  解析 {len(parsed["items"])} 筆, 跳過 停產 {parsed["skipped"]["discontinued"]} + 空白 {parsed["skipped"]["empty"]}')
            total += import_to_firestore(parsed, dry_run=args.dry_run)
        except Exception as e:
            print(f'  ❌ 失敗: {e}')

    print(f'\n總計: {total} 筆')


if __name__ == '__main__':
    main()
